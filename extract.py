import openpyxl, json, collections, datetime, warnings
warnings.filterwarnings('ignore')
wb = openpyxl.load_workbook('/sessions/festive-loving-thompson/mnt/uploads/RMA Sales and Expenses v2026.xlsx', data_only=True)

def d(x): return x.strftime('%Y-%m-%d') if isinstance(x, datetime.datetime) else None
def t(x): return x.strftime('%H:%M') if isinstance(x, datetime.time) else str(x) if x else None

# Sales Entry = main ledger
sales=[]
for r in wb['Sales Entry'].iter_rows(min_row=2, values_only=True):
    if r[0] is None or not isinstance(r[1], datetime.datetime): continue
    sales.append({'id':r[0],'date':d(r[1]),'client':r[2],'slot':t(r[3]),'type':'Peak' if r[4] and 'Peak' in str(r[4]) else 'Off','amount':round((r[8] or 0)+(r[6] or 0),2),'paid':round(r[8] or 0,2)})

txns=[]
for r in wb['Transaction Entry'].iter_rows(min_row=2, values_only=True):
    if not isinstance(r[0], datetime.datetime): continue
    txns.append({'date':d(r[0]),'type':r[1],'cat':(r[2] or '').strip(),'details':r[3],'amount':round(r[5] or 0,2)})

json.dump({'sales':sales,'txns':txns}, open('/sessions/festive-loving-thompson/mnt/outputs/data.json','w'))
print(len(sales), len(txns))

# quick analytics for insights
from collections import Counter, defaultdict
dow = defaultdict(lambda: [0,0.0])   # weekday -> [count, revenue]
slot = defaultdict(lambda: [0,0.0])
month = defaultdict(lambda: [0,0.0])
clients = defaultdict(lambda: [0,0.0,None,None])
for s in sales:
    dt = datetime.date.fromisoformat(s['date'])
    dow[dt.strftime('%a')][0]+=1; dow[dt.strftime('%a')][1]+=s['paid']
    slot[s['slot']][0]+=1; slot[s['slot']][1]+=s['paid']
    m=s['date'][:7]; month[m][0]+=1; month[m][1]+=s['paid']
    c=clients[s['client']]; c[0]+=1; c[1]+=s['paid']
    c[2]=min(c[2] or s['date'], s['date']); c[3]=max(c[3] or s['date'], s['date'])
print('DOW:', {k:v for k,v in sorted(dow.items())})
print('SLOT:', dict(sorted(slot.items())))
mm = sorted(month.items())
print('last 14 months:'); [print(' ',k,v) for k,v in mm[-14:]]
top = sorted(clients.items(), key=lambda x:-x[1][1])[:12]
print('TOP CLIENTS:'); [print(' ',k,v) for k,v in top]
print('unique clients:', len(clients))
# retention: clients whose last booking > 60 days before 2026-06-30 but had >=5 bookings
lapsed=[(k,v) for k,v in clients.items() if v[0]>=5 and v[3]<'2026-05-01']
print('lapsed regulars (>=5 bookings, none since May 2026):', len(lapsed))
# avg price
peak=[s['paid'] for s in sales if s['type']=='Peak']; off=[s['paid'] for s in sales if s['type']=='Off']
print('avg peak:', sum(peak)/len(peak), 'avg off:', sum(off)/len(off))
# occupancy: 11 slots/day
first=min(s['date'] for s in sales); last=max(s['date'] for s in sales)
days=(datetime.date.fromisoformat(last)-datetime.date.fromisoformat(first)).days+1
print('overall occupancy:', len(sales)/(days*11))
# 2026 YTD occupancy
s26=[s for s in sales if s['date']>='2026-01-01']
days26=(datetime.date.fromisoformat(last)-datetime.date(2026,1,1)).days+1
print('2026 occupancy:', len(s26)/(days26*11), 'bookings:', len(s26))
